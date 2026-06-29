from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet


ADD_FILL_COLOR = "ADD8E6"
DELETE_FILL_COLOR = "FF9393"
SUMMARY_SHEET_NAME = "summary_table"
COMMENT_AUTHOR = "DocTools API"


def normalize_cell_value(cell_value: Any) -> str:
    if cell_value is None:
        return ""

    return str(cell_value).strip()


def cell_has_content(cell_value: Any) -> bool:
    return normalize_cell_value(cell_value) != ""


def build_change_comment(change_type: str, changed_value: Any) -> Comment:
    comment_text = f"{change_type}: {normalize_cell_value(changed_value)}"
    return Comment(comment_text, COMMENT_AUTHOR)


def apply_cell_change_style(
    cell: Cell,
    fill_color: str,
    change_type: str,
    changed_value: Any,
) -> None:
    cell.fill = PatternFill(
        fill_type="solid",
        fgColor=fill_color,
    )
    cell.comment = build_change_comment(change_type, changed_value)


def copy_cell_value_and_style(source_cell: Cell, target_cell: Cell) -> None:
    target_cell.value = source_cell.value

    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)

    if source_cell.hyperlink:
        target_cell.hyperlink = source_cell.hyperlink.target

    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)


def get_cell_display_value(cell: Cell) -> Any:
    return cell.value


def get_worksheet_max_row(worksheet: Worksheet | None) -> int:
    if worksheet is None:
        return 0

    return worksheet.max_row


def get_worksheet_max_column(worksheet: Worksheet | None) -> int:
    if worksheet is None:
        return 0

    return worksheet.max_column


def get_comparison_limit(
    original_worksheet: Worksheet | None,
    modified_worksheet: Worksheet | None,
) -> tuple[int, int]:
    max_row = max(
        get_worksheet_max_row(original_worksheet),
        get_worksheet_max_row(modified_worksheet),
    )
    max_column = max(
        get_worksheet_max_column(original_worksheet),
        get_worksheet_max_column(modified_worksheet),
    )

    return max_row, max_column


def get_worksheet_by_name(workbook, sheet_name: str) -> Worksheet | None:
    if sheet_name not in workbook.sheetnames:
        return None

    return workbook[sheet_name]


def get_compared_worksheet(workbook, sheet_name: str) -> Worksheet:
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]

    return workbook.create_sheet(sheet_name)


def clear_existing_summary_sheet(workbook) -> None:
    if SUMMARY_SHEET_NAME not in workbook.sheetnames:
        return

    summary_worksheet = workbook[SUMMARY_SHEET_NAME]
    workbook.remove(summary_worksheet)


def build_ordered_sheet_names(original_workbook, modified_workbook) -> list[str]:
    ordered_sheet_names = list(modified_workbook.sheetnames)

    for sheet_name in original_workbook.sheetnames:
        if sheet_name not in ordered_sheet_names:
            ordered_sheet_names.append(sheet_name)

    if SUMMARY_SHEET_NAME in ordered_sheet_names:
        ordered_sheet_names.remove(SUMMARY_SHEET_NAME)

    return ordered_sheet_names


def mark_added_cell(compared_cell: Cell, modified_cell: Cell) -> int:
    if not cell_has_content(modified_cell.value):
        return 0

    apply_cell_change_style(
        compared_cell,
        ADD_FILL_COLOR,
        "add",
        get_cell_display_value(modified_cell),
    )

    return 1


def mark_deleted_cell(compared_worksheet: Worksheet, original_cell: Cell) -> int:
    if not cell_has_content(original_cell.value):
        return 0

    compared_cell = compared_worksheet.cell(
        row=original_cell.row,
        column=original_cell.column,
    )

    copy_cell_value_and_style(original_cell, compared_cell)

    apply_cell_change_style(
        compared_cell,
        DELETE_FILL_COLOR,
        "delete",
        get_cell_display_value(original_cell),
    )

    return 1


def mark_replaced_cell(
    compared_cell: Cell,
    original_cell: Cell,
    modified_cell: Cell,
) -> tuple[int, int]:
    if not cell_has_content(original_cell.value) and cell_has_content(modified_cell.value):
        added_count = mark_added_cell(compared_cell, modified_cell)
        return added_count, 0

    if cell_has_content(original_cell.value) and not cell_has_content(modified_cell.value):
        apply_cell_change_style(
            compared_cell,
            DELETE_FILL_COLOR,
            "delete",
            get_cell_display_value(original_cell),
        )
        return 0, 1

    apply_cell_change_style(
        compared_cell,
        ADD_FILL_COLOR,
        "add",
        get_cell_display_value(modified_cell),
    )

    previous_value_comment = build_change_comment(
        "delete",
        get_cell_display_value(original_cell),
    )

    compared_cell.comment = Comment(
        f"{previous_value_comment.text}\nadd: {normalize_cell_value(modified_cell.value)}",
        COMMENT_AUTHOR,
    )

    return 1, 1


def compare_existing_worksheet_cells(
    original_worksheet: Worksheet,
    modified_worksheet: Worksheet,
    compared_worksheet: Worksheet,
) -> tuple[int, int]:
    total_additions = 0
    total_deletions = 0

    max_row, max_column = get_comparison_limit(
        original_worksheet,
        modified_worksheet,
    )

    for row_index in range(1, max_row + 1):
        for column_index in range(1, max_column + 1):
            original_cell = original_worksheet.cell(row=row_index, column=column_index)
            modified_cell = modified_worksheet.cell(row=row_index, column=column_index)
            compared_cell = compared_worksheet.cell(row=row_index, column=column_index)

            original_value = normalize_cell_value(original_cell.value)
            modified_value = normalize_cell_value(modified_cell.value)

            if original_value == modified_value:
                continue

            if not cell_has_content(original_cell.value):
                total_additions += mark_added_cell(compared_cell, modified_cell)
                continue

            if not cell_has_content(modified_cell.value):
                total_deletions += mark_deleted_cell(compared_worksheet, original_cell)
                continue

            added_count, deleted_count = mark_replaced_cell(
                compared_cell,
                original_cell,
                modified_cell,
            )

            total_additions += added_count
            total_deletions += deleted_count

    return total_additions, total_deletions


def mark_added_worksheet(compared_worksheet: Worksheet) -> int:
    total_additions = 0

    for row_cells in compared_worksheet.iter_rows():
        for compared_cell in row_cells:
            total_additions += mark_added_cell(compared_cell, compared_cell)

    return total_additions


def mark_deleted_worksheet(
    compared_workbook,
    original_worksheet: Worksheet,
) -> int:
    compared_worksheet = get_compared_worksheet(
        compared_workbook,
        original_worksheet.title,
    )

    total_deletions = 0

    for row_cells in original_worksheet.iter_rows():
        for original_cell in row_cells:
            total_deletions += mark_deleted_cell(
                compared_worksheet,
                original_cell,
            )

    return total_deletions


def compare_workbook_sheets(
    original_workbook,
    modified_workbook,
    compared_workbook,
) -> tuple[int, int]:
    total_additions = 0
    total_deletions = 0

    ordered_sheet_names = build_ordered_sheet_names(
        original_workbook,
        modified_workbook,
    )

    for sheet_name in ordered_sheet_names:
        original_worksheet = get_worksheet_by_name(original_workbook, sheet_name)
        modified_worksheet = get_worksheet_by_name(modified_workbook, sheet_name)
        compared_worksheet = get_compared_worksheet(compared_workbook, sheet_name)

        if original_worksheet is None and modified_worksheet is not None:
            total_additions += mark_added_worksheet(compared_worksheet)
            continue

        if original_worksheet is not None and modified_worksheet is None:
            total_deletions += mark_deleted_worksheet(
                compared_workbook,
                original_worksheet,
            )
            continue

        if original_worksheet is None or modified_worksheet is None:
            continue

        added_count, deleted_count = compare_existing_worksheet_cells(
            original_worksheet,
            modified_worksheet,
            compared_worksheet,
        )

        total_additions += added_count
        total_deletions += deleted_count

    return total_additions, total_deletions


def build_summary_table(total_additions: int, total_deletions: int) -> dict[str, int]:
    return {
        "add": total_additions,
        "delete": total_deletions,
        "total_changes": total_additions + total_deletions,
    }


def add_summary_sheet(workbook, summary_table: dict[str, int]) -> None:
    clear_existing_summary_sheet(workbook)

    summary_worksheet = workbook.create_sheet(SUMMARY_SHEET_NAME)
    summary_rows = [
        ("add", summary_table["add"]),
        ("delete", summary_table["delete"]),
        ("total_changes", summary_table["total_changes"]),
    ]

    for row_index, summary_row in enumerate(summary_rows, start=1):
        summary_worksheet.cell(row=row_index, column=1).value = summary_row[0]
        summary_worksheet.cell(row=row_index, column=2).value = summary_row[1]


def save_processed_workbook(workbook, processed_path: Path) -> None:
    processed_path.parent.mkdir(parents=True, exist_ok=True)

    if processed_path.exists():
        processed_path.unlink()

    workbook.save(processed_path)


def compare_files(
    original_path: Path,
    modified_path: Path,
    processed_path: Path,
) -> dict[str, object]:
    """
    Compara dois arquivos XLSX e gera um terceiro arquivo baseado no modificado.

    Args:
        original_path: caminho do arquivo Excel original.
        modified_path: caminho do arquivo Excel modificado.
        processed_path: caminho onde o arquivo comparado será salvo.

    Returns:
        Dicionário com a tabela de resumo da comparação.
    """
    print("iniciando comparação de documentos Excel")

    try:
        original_workbook = load_workbook(original_path, data_only=False)
        modified_workbook = load_workbook(modified_path, data_only=False)
        compared_workbook = load_workbook(modified_path, data_only=False)

        total_additions, total_deletions = compare_workbook_sheets(
            original_workbook,
            modified_workbook,
            compared_workbook,
        )

        summary_table = build_summary_table(
            total_additions,
            total_deletions,
        )

        add_summary_sheet(compared_workbook, summary_table)
        save_processed_workbook(compared_workbook, processed_path)

        print("comparação de documentos Excel concluída com sucesso")

        return {
            "summary_table": summary_table,
        }

    except Exception as error_message:
        print(f"erro ao comparar documentos Excel: {error_message}")
        raise